import sqlite3

from xlrd import open_workbook

from itertools import zip_longest

from generate_file import delete_Blanks
import openpyxl
#Remove spaces, uppercase all the letters
def deal_string(words):
    a = words.replace(' ','')
    b = a.upper()
    return b

def excel_to_dict(name_of_file_excel):
    book = open_workbook(name_of_file_excel, formatting_info=False)

    cursheet = book.sheet_by_index(0)

    map_result = list()
    for row in range(cursheet.nrows):
        if row == 0:
            colum_list = [cursheet.cell(0, col).value for col in range(cursheet.ncols)]
        else:
            content = [cursheet.cell(row, col).value for col in range(cursheet.ncols)]
            content_map = map(lambda x: x.strip() if isinstance(x, str) else x, content)
            map_result.append(dict(zip_longest(colum_list, content_map)))
    return map_result

def sql_phrase_create_table(keys):
    phrase1 = ""
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase1 = keys[i]+" CHAR(30) NOT NULL"
        else:
            phrase1 = phrase1 + "," + keys[i]+" CHAR(30) NOT NULL"
    return phrase1

def sql_phrase_insert_values(keys, dictionary, cate):
    phrase1 = ""
    phrase2 = ""
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase2 = "'" + str(dictionary[keys[i]])
        elif i == l-1:
            phrase2 = phrase2 + "','" +str(dictionary[keys[i]]) + "'"
        else:
            phrase2 = phrase2 + "','" +str(dictionary[keys[i]])
    for j in range(l):
        if j == 0:
            phrase1 = keys[j]
        else:
            phrase1 = phrase1 + "," + keys[j]
    phrase3 = "INSERT INTO "+cate+" (" +phrase1+ ") VALUES ("+phrase2+")"
    return phrase3

def delete_duplicate(name_table):
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute("SELECT sql FROM sqlite_master WHERE type = 'table' AND name = '"+ name_table +"'")
    ress = cur.fetchall()
    phrase = ress[0][0]
    phrase = str(phrase)
    phrase = phrase.replace(name_table, "TMP", 1)
    cur.execute(phrase)
    cur.execute("insert into TMP select distinct * from "+name_table)
    cur.execute("drop table "+ name_table)
    cur.execute("SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'TMP'")
    ress = cur.fetchall()
    phrase = ress[0][0]
    phrase = str(phrase)
    phrase = phrase.replace("TMP", name_table, 1)
    cur.execute(phrase)
    cur.execute("insert into "+name_table+" select distinct * from TMP")
    cur.execute("drop table TMP")
    cur.close()
    conn.commit()
    conn.close()

def test_change_p():
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute("SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'DOOR'")
    ress = cur.fetchall()
    phrase = ress[0][0]
    phrase = str(phrase)
    phrase = phrase.replace("DOOR", "TMP", 1)
    print(phrase)
    cur.execute(phrase)
    cur.close()
    conn.commit()
    conn.close()


def generate_database(name_of_file_excel, name_table):
    conn = sqlite3.connect("database.db")
    list_dict = excel_to_dict(name_of_file_excel)
    ks = list(list_dict[0].keys())
    p1 = sql_phrase_create_table(ks);
    p1 = "create table if not exists "+name_table+" (" +p1+ ")"
    cursor = conn.cursor()
    cursor.execute(p1)
    for i in range(len(list_dict)):
        dictx = list_dict[i]
        dicty = delete_Blanks(dictx)
        keys = list(dicty.keys())
        p2 = sql_phrase_insert_values(keys, dicty, name_table)
        cursor.execute(p2)
    cursor.close()
    conn.commit()
    conn.close()
    delete_duplicate(name_table)

def delete_table(name_table):
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute("drop table "+ name_table)
    cur.close()
    conn.commit()
    conn.close()

def print_table(name_table):
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute("select * from "+name_table)
    name_list = [tuple[0] for tuple in cur.description]
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(len(name_list)):
        ws.cell(row = 1, column = i+1).value = name_list[i]
    ress = cur.fetchall()
    for i in range(len(ress)):
        for j in range(len(ress[0])):
            ws.cell(row = i+2, column = j+1).value = ress[i][j]
    wb.save("print_database_"+name_table+".xlsx")
