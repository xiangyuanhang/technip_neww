import openpyxl
import os
import sqlite3
from xlrd import open_workbook
from openpyxl.styles import PatternFill
from itertools import zip_longest
import csv

#
def deal_string(words):
    a = words.replace(' ','')
    b = a.upper()
    return b
def excel_to_dict(name_of_file_excel):
    book = open_workbook(name_of_file_excel, formatting_info=False)
    cursheet = book.sheet_by_index(0)
    map_result = list()
    for row in range(cursheet.nrows):
        if row == 0:
            colum_list = [cursheet.cell(0, col).value for col in range(cursheet.ncols)]
        else:
            content = [cursheet.cell(row, col).value for col in range(cursheet.ncols)]
            content_map = map(lambda x: x.strip() if isinstance(x, str) else x, content)
            map_result.append(dict(zip_longest(colum_list, content_map)))
    return map_result
def delete_Blanks(array):
    tempArray = array.copy()
    for key, value in sorted(tempArray.items()):
        if value == "":
            del tempArray[key]
    return tempArray

def find_Blanks(array):
    liste = []
    tempArray = array.copy()
    for key, value in sorted(tempArray.items()):
        if value == "":
            liste.append(key)
    return liste

def input_rw(rwm):
    d = {}
    d[u'RW'] = rwm
    return d

def input_rw_sp(rwm, spm):
    d = {}
    d[u'RW'] = rwm
    d[u'SUPPLIER'] = spm
    return d

def merge_dict(dict1, dict2):
    dictmerge = dict(list(dict1.items())+list(dict2.items()))
    return dictmerge

def delete_duplicate(name_table):
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute("create table TMP as select distinct * from "+name_table)
    cur.execute("drop table "+ name_table)
    cur.execute("create table "+name_table+" as select * from TMP")
    cur.execute("drop table TMP")
    cur.close()
    conn.commit()
    conn.close()

def dict_to_csv(list_dict, file_name):
    dict1 = list_dict[0]
    k = dict1.keys()
    list_values = []
    for i in range(len(list_dict)):
        v = list_dict[i].values()
        list_values.append(v)
    with open(file_name,"w") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(k) # write colunm names
        writer.writerows(list_values)

fill0 = PatternFill("solid", fgColor="FFDEAD")
fill1 = PatternFill("solid", fgColor="C7C7C7")
fill2 = PatternFill("solid", fgColor="EED2EE")
fill3 = PatternFill("solid", fgColor="B3EE3A")

# for door
def sql_phrase_select_door(keys, dictionary):
    phrase = "select WEIGHT, RW from DOOR where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase+keys[i]+" = '"+str(dictionary[keys[i]])+"'"
        else:
            phrase = phrase+" and "+keys[i]+" = '"+str(dictionary[keys[i]])+"'"
    return phrase
def sql_phrase_select_supplier_door(keys, dictionary):
    phrase = "select SUPPLIER from DOOR where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase+keys[i]+" = '"+str(dictionary[keys[i]])+"'"
        else:
            phrase = phrase+" and "+keys[i]+" = '"+str(dictionary[keys[i]])+"'"
    return phrase
def find_supplier_door(dict_door,database):
    dict_door_new = delete_Blanks(dict_door)
    k = list(dict_door_new.keys())
    k.remove("HEIGHT")
    k.remove("WIDTH")
    k.remove("MARK")
    k.remove("LEVEL")
    p = sql_phrase_select_supplier_door(k, dict_door)
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    return ress
def check_key_values_door(d):
    l = ["FIRERATING","HEIGHT","WIDTH","FUNCTION", "LEVEL"]
    for i in range(len(l)):
        if d[l[i]] == "":
            return 1
    return 0
def find_weight_door(dict_door,database):
    dict_door_new = delete_Blanks(dict_door)
    k = list(dict_door_new.keys())
    k.remove("WIDTH")
    k.remove("HEIGHT")
    k.remove("MARK")
    k.remove("LEVEL")
    p = sql_phrase_select_door(k, dict_door)
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    if len(ress) == 1:
        return (float(ress[0][0])*float(dict_door["WIDTH"])*float(dict_door["HEIGHT"])/1000000.0,"d", "0")
    elif len(ress) > 1:
        return ("duo",ress)
    else:
        return (0,"b")
def sql_phrase_select_door_final(keys, dictionary):
    phrase = "select WEIGHT from DOOR where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase+keys[i]+" = '"+str(dictionary[keys[i]])+"'"
        else:
            phrase = phrase+" and "+keys[i]+" = '"+str(dictionary[keys[i]])+"'"
    return phrase
def find_weight_level_door_final(dict_door):
    dict_door_new = delete_Blanks(dict_door)
    k = list(dict_door_new.keys())
    k.remove("HEIGHT")
    k.remove("WIDTH")
    k.remove("MARK")
    k.remove("LEVEL")
    p = sql_phrase_select_door_final(k, dict_door)
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    if len(ress)==1:
        return (float(ress[0][0])*float(dict_door["WIDTH"])*float(dict_door["HEIGHT"])/1000000.0, dict_door["LEVEL"])
    else:
        return (0.0,0.0)
def insert_weight_level_door(dict_door):
    weight_level_door = find_weight_level_door_final(dict_door)
    phrase = "insert into WEIGHT_DOOR (WEIGHT, LEVEL) values ('"+str(weight_level_door[0])+"', '"+str(weight_level_door[1])+"')"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(phrase)
    cur.close()
    conn.commit()
    conn.close()
def create_table_weight_door():
    pp = "drop table WEIGHT_DOOR"
    p = "create table if not exists WEIGHT_DOOR (WEIGHT CHAR(30), LEVEL CHAR(30))"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    cur.execute(pp)
    cur.execute(p)
    cur.close()
    conn.commit()
    conn.close()
def group_weight_door():
    p = "select LEVEL, sum(WEIGHT) from WEIGHT_DOOR group by LEVEL"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    return ress
def write_weight_door(ress):
    number_row = len(ress)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "summary"
    ws["A1"] = "SUMMARY"
    ws["A2"] = "ELEMENT"
    ws["B2"] = "WEIGHT"
    ws["C2"] = "UNIT"
    for i in range(number_row):
        ws["A" + str(i + 3)] = "DOOR "+str(ress[i][0])
        ws["A" + str(i + 3)].fill = fill0
        ws["B" + str(i + 3)] = str(round(ress[i][1],4))
        ws["B" + str(i + 3)].fill = fill0
        ws["C" + str(i + 3)] = "kg"
        ws["C" + str(i + 3)].fill = fill0
    wb.save("weight_report.xlsx")
def sql_phrase_insert_3_values_door(keys, dictionary):
    phrase = "insert into QUANTITY_DOOR (FUNCTION, FIRERATING, SINGLEDOUBLE) select FUNCTION, FIRERATING, SINGLEDOUBLE from DOOR where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
        else:
            phrase = phrase + " and " + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
    return phrase
def insert_3_values_door(dict_door):
    dict_door_new = delete_Blanks(dict_door)
    k = list(dict_door_new.keys())
    k.remove("HEIGHT")
    k.remove("WIDTH")
    k.remove("MARK")
    k.remove("LEVEL")
    p = sql_phrase_insert_3_values_door(k, dict_door)
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    cur.close()
    conn.commit()
    conn.close()
def create_table_quantity_door():
    pp = "drop table QUANTITY_DOOR"
    p = "create table if not exists QUANTITY_DOOR (FUNCTION CHAR(30), FIRERATING CHAR(30), SINGLEDOUBLE CHAR(30))"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    cur.execute(pp)
    cur.execute(p)
    cur.close()
    conn.commit()
    conn.close()
def group_quantity_door():
    p = "select FUNCTION, FIRERATING, SINGLEDOUBLE, count(*) from QUANTITY_DOOR group by FUNCTION, FIRERATING, SINGLEDOUBLE"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    return ress
def write_quantity_door(ress):
    number_row = len(ress)
    if os.path.isfile("price_schedule.xlsx"):
        wb = openpyxl.load_workbook("price_schedule.xlsx")
        names = wb.get_sheet_names()
        if "result_door" in names:
            ws = wb["result_door"]
            ws["A1"] = "summary"
            ws["A2"] = "FUNCTION"
            ws["B2"] = "FIRERATING"
            ws["C2"] = "SINGLEDOUBLE"
            ws["D2"] = "COUNT"
            for i in range(number_row):
                ws["A"+str(i+3)] = str(ress[i][0])
                ws["B" + str(i + 3)] = str(ress[i][1])
                ws["C" + str(i + 3)] = str(ress[i][2])
                ws["D" + str(i + 3)] = str(ress[i][3])
            wb.save("price_schedule.xlsx")
        else:
            ws = wb.create_sheet("result_door")
            ws["A1"] = "summary"
            ws["A2"] = "FUNCTION"
            ws["B2"] = "FIRERATING"
            ws["C2"] = "SINGLEDOUBLE"
            ws["D2"] = "COUNT"
            for i in range(number_row):
                ws["A" + str(i + 3)] = str(ress[i][0])
                ws["B" + str(i + 3)] = str(ress[i][1])
                ws["C" + str(i + 3)] = str(ress[i][2])
                ws["D" + str(i + 3)] = str(ress[i][3])
            wb.save("price_schedule.xlsx")
    else:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("result_door")
        ws["A1"] = "summary"
        ws["A2"] = "FUNCTION"
        ws["B2"] = "FIRERATING"
        ws["C2"] = "SINGLEDOUBLE"
        ws["D2"] = "COUNT"
        for i in range(number_row):
            ws["A" + str(i + 3)] = str(ress[i][0])
            ws["B" + str(i + 3)] = str(ress[i][1])
            ws["C" + str(i + 3)] = str(ress[i][2])
            ws["D" + str(i + 3)] = str(ress[i][3])
        wb.save("price_schedule.xlsx")

#for floor
def sql_phrase_select_floor(keys, dictionary):
    phrase = "select WEIGHT, RW from FLOOR where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
        else:
            phrase = phrase + " and " + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
    return phrase
def sql_phrase_select_supplier_floor(keys, dictionary):
    phrase = "select SUPPLIER from FLOOR where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
        else:
            phrase = phrase + " and " + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
    return phrase
def find_supplier_floor(dict_floor, database):
    dict_floor_new = delete_Blanks(dict_floor)
    k = list(dict_floor_new.keys())
    k.remove("MATERIALAREA")
    k.remove("LEVEL")
    k.remove("ELEVATIONATTOP")
    p = sql_phrase_select_supplier_floor(k, dict_floor)
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    return ress
def check_key_values_floor(d):
    l = ["FIRERATING", "MATERIALAREA", "LEVEL"]
    for i in range(len(l)):
        if d[l[i]] == "":
            return 1
    return 0
def find_weight_floor(dict_floor,database):
    dict_floor_new = delete_Blanks(dict_floor)
    k = list(dict_floor_new.keys())
    k.remove("MATERIALAREA")
    k.remove("LEVEL")
    k.remove("ELEVATIONATTOP")
    p = sql_phrase_select_floor(k, dict_floor)
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    if len(ress) == 1:
        return (float(ress[0][0])*float(dict_floor["MATERIALAREA"]),"d", dict_floor["MATERIALAREA"])
    elif len(ress) > 1:
        return ("duo",ress)
    else:
        return (0,"b")
def group_weight_floor():
    p = "select LEVEL, sum(WEIGHT) from WEIGHT_FLOOR group by LEVEL"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    return ress
def create_table_quantity_floor():
    pp = "drop table QUANTITY_FLOOR"
    p = "create table if not exists QUANTITY_FLOOR (TYPE CHAR(30), MATERIALAREA CHAR(30))"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    cur.execute(pp)
    cur.execute(p)
    cur.close()
    conn.commit()
    conn.close()
def create_table_weight_floor():
    pp = "drop table WEIGHT_FLOOR"
    p = "create table if not exists WEIGHT_FLOOR (LEVEL CHAR(30), WEIGHT CHAR(30))"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    cur.execute(pp)
    cur.execute(p)
    cur.close()
    conn.commit()
    conn.close()
def sql_phrase_select_floor_final(keys, dictionary):
    phrase = "select WEIGHT from FLOOR where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
        else:
            phrase = phrase + " and " + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
    return phrase
def find_weight_level_floor_final(dict_floor):
    k = list(dict_floor.keys())
    k.remove("MATERIALAREA")
    k.remove("LEVEL")
    k.remove("ELEVATIONATTOP")
    p = sql_phrase_select_floor_final(k, dict_floor)
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    if len(ress) == 1:
        return (float(ress[0][0]) * float(dict_floor["MATERIALAREA"]), dict_floor["LEVEL"])
    else:
        return (0.0, 0.0)
def insert_weight_level_floor(dict_floor):
    weight_level_floor = find_weight_level_floor_final(dict_floor)
    phrase = "insert into WEIGHT_FLOOR (WEIGHT, LEVEL) values ('"+str(weight_level_floor[0])+"', '"+str(weight_level_floor[1])+"')"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(phrase)
    cur.close()
    conn.commit()
    conn.close()
def insert_type_area_floor(dict_floor):
    p = "insert into QUANTITY_FLOOR(TYPE, MATERIALAREA) values ('"+str(dict_floor["TYPE"])+"', '"+str(dict_floor["MATERIALAREA"])+"')"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    cur.close()
    conn.commit()
    conn.close()
def group_quantity_floor():
    p = "select TYPE, sum(MATERIALAREA) from QUANTITY_FLOOR group by TYPE"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    return ress

def write_quantity_floor(ress):
    number_row = len(ress)
    if os.path.isfile("price_schedule.xlsx"):
        wb = openpyxl.load_workbook("price_schedule.xlsx")
        names = wb.get_sheet_names()
        if "result_floor" in names:
            ws = wb["result_floor"]
            ws["A1"] = "summary"
            ws["A2"] = "TYPE"
            ws["B2"] = "QUANTITY"
            ws["C2"] = "UNIT"
            for i in range(number_row):
                ws["A"+str(i+3)] = str(ress[i][0])
                ws["B" + str(i + 3)] = str(round(ress[i][1],4))
                ws["C" + str(i + 3)] = "m²"
            wb.save("price_schedule.xlsx")
        else:
            ws = wb.create_sheet("result_floor")
            ws["A1"] = "summary"
            ws["A2"] = "TYPE"
            ws["B2"] = "QUANTITY"
            ws["C2"] = "UNIT"
            for i in range(number_row):
                ws["A" + str(i + 3)] = str(ress[i][0])
                ws["B" + str(i + 3)] = str(round(ress[i][1], 4))
                ws["C" + str(i + 3)] = "m²"
            wb.save("price_schedule.xlsx")
    else:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("result_floor")
        ws["A1"] = "summary"
        ws["A2"] = "TYPE"
        ws["B2"] = "QUANTITY"
        ws["C2"] = "UNIT"
        for i in range(number_row):
            ws["A" + str(i + 3)] = str(ress[i][0])
            ws["B" + str(i + 3)] = str(round(ress[i][1], 4))
            ws["C" + str(i + 3)] = "m²"
        wb.save("price_schedule.xlsx")

def write_weight_floor(ress):
    number_row = len(ress)
    if os.path.isfile("weight_report.xlsx"):
        wb = openpyxl.load_workbook("weight_report.xlsx")
        ws = wb.get_sheet_by_name("summary")
        nr = ws.max_row
        for i in range(number_row):
            ws["A" + str(i + 1 + nr)] = "FLOOR "+str(ress[i][0])
            ws["A" + str(i + 1 + nr)].fill = fill1
            ws["B" + str(i + 1 + nr)] = str(round(ress[i][1],4))
            ws["B" + str(i + 1 + nr)].fill = fill1
            ws["C" + str(i + 1 + nr)] = "kg"
            ws["C" + str(i + 1 + nr)].fill = fill1
        wb.save("weight_report.xlsx")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "summary"
        ws["A1"] = "SUMMARY"
        ws["A2"] = "ELEMENT"
        ws["B2"] = "WEIGHT"
        ws["C2"] = "UNIT"
        for i in range(number_row):
            ws["A" + str(i + 3)] = "FLOOR "+str(ress[i][0])
            ws["A" + str(i + 3)].fill = fill1
            ws["B" + str(i + 3)] = str(round(ress[i][1],4))
            ws["B" + str(i + 3)].fill = fill1
            ws["C" + str(i + 3)] = "kg"
            ws["C" + str(i + 3)].fill = fill1
        wb.save("weight_report.xlsx")

#for insulation
def sql_phrase_select_insulation(keys, dictionary):
    phrase = "select DENSITY, RW from INSULATION where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
        else:
            phrase = phrase + " and " + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
    return phrase
def sql_phrase_select_supplier_insulation(keys, dictionary):
    phrase = "select SUPPLIER from INSULATION where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
        else:
            phrase = phrase + " and " + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
    return phrase
def find_supplier_insulation(dict_insulation, database):
    dict_insulation_new = delete_Blanks(dict_insulation)
    k = list(dict_insulation_new.keys())
    k.remove("MATERIALVOLUME")
    k.remove("LEVEL")
    k.remove("MARK")
    p = sql_phrase_select_supplier_insulation(k, dict_insulation_new)
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    return ress
def check_key_values_insulation(d):
    l = ["FIRERATING", "MATERIALVOLUME", "LEVEL"]
    for i in range(len(l)):
        if d[l[i]] == "":
            return 1
    return 0
def find_weight_insulation(dict_insulation,database):
    dict_insulation_new = delete_Blanks(dict_insulation)
    k = list(dict_insulation_new.keys())
    k.remove("MATERIALVOLUME")
    k.remove("LEVEL")
    k.remove("MARK")
    p = sql_phrase_select_insulation(k, dict_insulation_new)
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    if len(ress) == 1:
        return (float(ress[0][0])*float(dict_insulation_new["MATERIALVOLUME"]),"d", dict_insulation_new["MATERIALVOLUME"])
    elif len(ress) > 1:
        return ("duo",ress)
    else:
        return (0,"b")
def create_table_quantity_insulation():
    pp = "drop table QUANTITY_INSULATION"
    p = "create table if not exists QUANTITY_INSULATION (FIRERATING CHAR(30), MATERIALVOLUME CHAR(30))"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    cur.execute(pp)
    cur.execute(p)
    cur.close()
    conn.commit()
    conn.close()
def insert_firerating_volume_insulation(dict_insulation):
    p = "insert into QUANTITY_INSULATION(FIRERATING, MATERIALVOLUME) values ('"+str(dict_insulation["FIRERATING"])+"', '"+str(dict_insulation["MATERIALVOLUME"])+"')"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    cur.close()
    conn.commit()
    conn.close()
def group_quantity_insulation():
    p = "select FIRERATING, sum(MATERIALVOLUME) from QUANTITY_INSULATION group by FIRERATING"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    return ress

def write_quantity_insulation(ress):
    number_row = len(ress)
    if os.path.isfile("price_schedule.xlsx"):
        wb = openpyxl.load_workbook("price_schedule.xlsx")
        names = wb.get_sheet_names()
        if "result_insulation" in names:
            ws = wb["result_insulation"]
            ws["A1"] = "summary"
            ws["A2"] = "FIRERATING"
            ws["B2"] = "QUANTITY"
            ws["C2"] = "UNIT"
            for i in range(number_row):
                ws["A"+str(i+3)] = str(ress[i][0])
                ws["B" + str(i + 3)] = str(round(ress[i][1],4))
                ws["C" + str(i + 3)] = "m³"
            wb.save("price_schedule.xlsx")
        else:
            ws = wb.create_sheet("result_insulation")
            ws["A1"] = "summary"
            ws["A2"] = "FIRERATING"
            ws["B2"] = "QUANTITY"
            ws["C2"] = "UNIT"
            for i in range(number_row):
                ws["A" + str(i + 3)] = str(ress[i][0])
                ws["B" + str(i + 3)] = str(round(ress[i][1], 4))
                ws["C" + str(i + 3)] = "m³"
            wb.save("price_schedule.xlsx")
    else:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("result_insulation")
        ws["A1"] = "summary"
        ws["A2"] = "FIRERATING"
        ws["B2"] = "QUANTITY"
        ws["C2"] = "UNIT"
        for i in range(number_row):
            ws["A" + str(i + 3)] = str(ress[i][0])
            ws["B" + str(i + 3)] = str(round(ress[i][1], 4))
            ws["C" + str(i + 3)] = "m³"
        wb.save("price_schedule.xlsx")
def create_table_weight_insulation():
    pp = "drop table WEIGHT_INSULATION"
    p = "create table if not exists WEIGHT_INSULATION (LEVEL CHAR(30), WEIGHT CHAR(30))"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    cur.execute(pp)
    cur.execute(p)
    cur.close()
    conn.commit()
    conn.close()
def sql_phrase_select_insulation_final(keys, dictionary):
    phrase = "select DENSITY from INSULATION where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
        else:
            phrase = phrase + " and " + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
    return phrase
def find_weight_level_insulation_final(dict_insulation):
    l = find_Blanks(dict_insulation)
    dict_insulation_new = delete_Blanks(dict_insulation)
    k = list(dict_insulation_new.keys())
    k.remove("MATERIALVOLUME")
    k.remove("LEVEL")
    k.remove("MARK")
    p = sql_phrase_select_insulation_final(k, dict_insulation_new)
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    if len(ress) == 1:
        return (float(ress[0][0]) * float(dict_insulation_new["MATERIALVOLUME"]), dict_insulation_new["LEVEL"])
    else:
        return (0.0, 0.0)
def insert_weight_level_insulation(dict_insulation):
    weight_level_insulation = find_weight_level_insulation_final(dict_insulation)
    phrase = "insert into WEIGHT_INSULATION (WEIGHT, LEVEL) values ('"+str(weight_level_insulation[0])+"', '"+str(weight_level_insulation[1])+"')"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(phrase)
    cur.close()
    conn.commit()
    conn.close()
def group_weight_insulation():
    p = "select LEVEL, sum(WEIGHT) from WEIGHT_INSULATION group by LEVEL"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    return ress
def write_weight_insulation(ress):
    number_row = len(ress)
    if os.path.isfile("weight_report.xlsx"):
        wb = openpyxl.load_workbook("weight_report.xlsx")
        ws = wb.get_sheet_by_name("summary")
        nr = ws.max_row
        for i in range(number_row):
            ws["A" + str(i + 1 + nr)] = "INSULATION "+str(ress[i][0])
            ws["A" + str(i + 1 + nr)].fill = fill2
            ws["B" + str(i + 1 + nr)] = str(round(ress[i][1],4))
            ws["B" + str(i + 1 + nr)].fill = fill2
            ws["C" + str(i + 1 + nr)] = "kg"
            ws["C" + str(i + 1 + nr)].fill = fill2
        wb.save("weight_report.xlsx")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "summary"
        ws["A1"] = "SUMMARY"
        ws["A2"] = "ELEMENT"
        ws["B2"] = "WEIGHT"
        ws["C2"] = "UNIT"
        for i in range(number_row):
            ws["A" + str(i + 3)] = "INSULATION "+str(ress[i][0])
            ws["A" + str(i + 3)].fill = fill2
            ws["B" + str(i + 3)] = str(round(ress[i][1],4))
            ws["B" + str(i + 3)].fill = fill2
            ws["C" + str(i + 3)] = "kg"
            ws["C" + str(i + 3)].fill = fill2
        wb.save("weight_report.xlsx")

#for partition
def sql_phrase_select_partition(keys, dictionary):
    phrase = "select WEIGHT, RW from PARTITION where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
        else:
            phrase = phrase + " and " + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
    return phrase

def sql_phrase_select_supplier_partition(keys, dictionary):
    phrase = "select SUPPLIER from PARTITION where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
        else:
            phrase = phrase + " and " + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
    return phrase




def find_supplier_partition(dict_partition, database):
    dict_partition_new = delete_Blanks(dict_partition)
    k = list(dict_partition_new.keys())
    k.remove("AREA")
    k.remove("LEVEL")
    k.remove("MARK")
    p = sql_phrase_select_supplier_partition(k, dict_partition)
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    return ress



def check_key_values_partition(d):
    l = ["FIRERATING", "AREA", "LEVEL"]
    for i in range(len(l)):
        if d[l[i]] == "":
            return 1
    return 0


def find_weight_partition(dict_partition,database):
    dict_partition_new = delete_Blanks(dict_partition)
    k = list(dict_partition_new.keys())
    k.remove("AREA")
    k.remove("LEVEL")
    k.remove("MARK")
    p = sql_phrase_select_partition(k, dict_partition)
    conn = sqlite3.connect(database)
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    if len(ress) == 1:
        return (float(ress[0][0])*float(dict_partition["AREA"]),"d", dict_partition["AREA"])
    elif len(ress) > 1:
        return ("duo",ress)
    else:
        return (0,"b")


def create_table_quantity_partition():
    pp = "drop table QUANTITY_PARTITION"
    p = "create table if not exists QUANTITY_PARTITION (FIRERATING CHAR(30), AREA CHAR(30))"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    cur.execute(pp)
    cur.execute(p)
    cur.close()
    conn.commit()
    conn.close()
def insert_firerating_area_partition(dict_partition):
    p = "insert into QUANTITY_PARTITION(FIRERATING, AREA) values ('"+str(dict_partition["FIRERATING"])+"', '"+str(dict_partition["AREA"])+"')"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    cur.close()
    conn.commit()
    conn.close()
def group_quantity_partition():
    p = "select FIRERATING, sum(AREA) from QUANTITY_PARTITION group by FIRERATING"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    return ress

def write_quantity_partition(ress):
    number_row = len(ress)
    if os.path.isfile("price_schedule.xlsx"):
        wb = openpyxl.load_workbook("price_schedule.xlsx")
        names = wb.get_sheet_names()
        if "result_partition" in names:
            ws = wb["result_partition"]
            ws["A1"] = "summary"
            ws["A2"] = "FIRERATING"
            ws["B2"] = "QUANTITY"
            ws["C2"] = "UNIT"
            for i in range(number_row):
                ws["A"+str(i+3)] = str(ress[i][0])
                ws["B" + str(i + 3)] = str(round(ress[i][1],4))
                ws["C" + str(i + 3)] = "m²"
            wb.save("price_schedule.xlsx")
        else:
            ws = wb.create_sheet("result_partition")
            ws["A1"] = "summary"
            ws["A2"] = "FIRERATING"
            ws["B2"] = "QUANTITY"
            ws["C2"] = "UNIT"
            for i in range(number_row):
                ws["A" + str(i + 3)] = str(ress[i][0])
                ws["B" + str(i + 3)] = str(round(ress[i][1], 4))
                ws["C" + str(i + 3)] = "m²"
            wb.save("price_schedule.xlsx")
    else:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("result_partition")
        ws["A1"] = "summary"
        ws["A2"] = "FIRERATING"
        ws["B2"] = "QUANTITY"
        ws["C2"] = "UNIT"
        for i in range(number_row):
            ws["A" + str(i + 3)] = str(ress[i][0])
            ws["B" + str(i + 3)] = str(round(ress[i][1], 4))
            ws["C" + str(i + 3)] = "m²"
        wb.save("price_schedule.xlsx")
def create_table_weight_partition():
    pp = "drop table WEIGHT_PARTITION"
    p = "create table if not exists WEIGHT_PARTITION (LEVEL CHAR(30), WEIGHT CHAR(30))"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    cur.execute(pp)
    cur.execute(p)
    cur.close()
    conn.commit()
    conn.close()
def sql_phrase_select_partition_final(keys, dictionary):
    phrase = "select WEIGHT from PARTITION where "
    l = len(keys)
    for i in range(l):
        if i == 0:
            phrase = phrase + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
        else:
            phrase = phrase + " and " + keys[i] + " = '" + str(dictionary[keys[i]]) + "'"
    return phrase
def find_weight_level_partition_final(dict_partition):
    l = find_Blanks(dict_partition)
    dict_partition_new = delete_Blanks(dict_partition)
    k = list(dict_partition_new.keys())
    k.remove("AREA")
    k.remove("LEVEL")
    k.remove("MARK")
    p = sql_phrase_select_partition_final(k, dict_partition_new)
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    if len(ress) == 1:
        return (float(ress[0][0]) * float(dict_partition["AREA"]), dict_partition["LEVEL"])
    else:
        return (0.0, 0.0)
def insert_weight_level_partition(dict_partition):
    weight_level_partition = find_weight_level_partition_final(dict_partition)
    phrase = "insert into WEIGHT_PARTITION (WEIGHT, LEVEL) values ('"+str(weight_level_partition[0])+"', '"+str(weight_level_partition[1])+"')"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(phrase)
    cur.close()
    conn.commit()
    conn.close()
def group_weight_partition():
    p = "select LEVEL, sum(WEIGHT) from WEIGHT_PARTITION group by LEVEL"
    conn = sqlite3.connect("database.db")
    cur = conn.cursor()
    cur.execute(p)
    ress = cur.fetchall()
    cur.close()
    conn.commit()
    conn.close()
    return ress

def write_weight_partition(ress):
    number_row = len(ress)
    if os.path.isfile("weight_report.xlsx"):
        wb = openpyxl.load_workbook("weight_report.xlsx")
        ws = wb.get_sheet_by_name("summary")
        nr = ws.max_row
        for i in range(number_row):
            ws["A" + str(i + 1 + nr)] = "PARTITION "+str(ress[i][0])
            ws["A" + str(i + 1 + nr)].fill = fill3
            ws["B" + str(i + 1 + nr)] = str(round(ress[i][1],4))
            ws["B" + str(i + 1 + nr)].fill = fill3
            ws["C" + str(i + 1 + nr)] = "kg"
            ws["C" + str(i + 1 + nr)].fill = fill3
        wb.save("weight_report.xlsx")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "summary"
        ws["A1"] = "SUMMARY"
        ws["A2"] = "ELEMENT"
        ws["B2"] = "WEIGHT"
        ws["C2"] = "UNIT"
        for i in range(number_row):
            ws["A" + str(i + 3)] = "PARTITION "+str(ress[i][0])
            ws["A" + str(i + 3)].fill = fill3
            ws["B" + str(i + 3)] = str(round(ress[i][1],4))
            ws["B" + str(i + 3)].fill = fill3
            ws["C" + str(i + 3)] = "kg"
            ws["C" + str(i + 3)].fill = fill3
        wb.save("weight_report.xlsx")
